import sqlite3
from fpdf import FPDF
import openpyxl
import os

DB_NAME = "users.db"

def export_to_pdf(username):
    file_path = f"{username}_report.pdf"

    print(f"Generating PDF at: {file_path}")

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)

    try:
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        c.execute("SELECT category, amount, description, date FROM expenses WHERE username = ?", (username,))
        expenses = c.fetchall()
        conn.close()

        pdf.cell(200, 10, txt=f"Expense Report for {username}", ln=True, align="C")
        pdf.ln(10)

        for category, amount, description, date in expenses:
            line = f"{date} | {category} | Rs.{amount} | {description}"
            pdf.cell(200, 10, txt=line.encode('latin-1', 'replace').decode('latin-1'), ln=True)

        pdf.output(file_path)
        print(f"PDF saved at: {file_path}")
        return file_path

    except Exception as e:
        print("Error in export_to_pdf:", e)
        return None

def export_to_excel(username):
    
    file_path = f"{username}_report.xlsx"

    print(f"Generating Excel at: {file_path}")

    try:
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        c.execute("SELECT category, amount, description, date FROM expenses WHERE username = ?", (username,))
        expenses = c.fetchall()
        conn.close()

        if not expenses:
            print("No expenses found for Excel export.")
            return None

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Expenses"
        ws.append(["Date", "Category", "Amount", "Description"])

        for row in expenses:
            ws.append([row[3], row[0], row[1], row[2]])

        wb.save(file_path)
        print(f"Excel saved at: {file_path}")
        return file_path

    except Exception as e:
        print("Error in export_to_excel:", e)
        return None
